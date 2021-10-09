from eth_account import Account
from openpyxl import Workbook
from web3 import Web3


def createNewETHWallet():
    wb = Workbook()
    ws = wb.active
    count = 0
    index = [0] * 10
    bagCount = 1
    while True:
       
        count = count + 1
        account = Account.create()
        privateKey = account._key_obj
        publicKey = privateKey.public_key
        address = publicKey.to_checksum_address()
        if address[len(address) - 5:len(address)] == "11111":
            print("privateKey:", privateKey)
            print("publicKey:", publicKey)
            print("address:", address)
            print("count:", count)
            if index[1] == 0:
                index[1] = 1
                ws.cell(bagCount, 1).value = address
                ws.cell(bagCount, 2).value = privateKey.to_hex()
                bagCount = bagCount + 1
        elif address[len(address) - 5:len(address)] == "22222":
            print("privateKey:", privateKey)
            print("publicKey:", publicKey)
            print("address:", address)
            print("count:", count)
            if index[2] == 0:
                index[2] = 1
                ws.cell(bagCount, 1).value = address
                ws.cell(bagCount, 2).value = privateKey.to_hex()
                bagCount = bagCount + 1
        elif address[len(address) - 5:len(address)] == "33333":
            print("privateKey:", privateKey)
            print("publicKey:", publicKey)
            print("address:", address)
            print("count:", count)
            if index[3] == 0:
                index[3] = 1
                ws.cell(bagCount, 1).value = address
                ws.cell(bagCount, 2).value = privateKey.to_hex()
                bagCount = bagCount + 1
        elif address[len(address) - 5:len(address)] == "44444":
            print("privateKey:", privateKey)
            print("publicKey:", publicKey)
            print("address:", address)
            print("count:", count)
            if index[4] == 0:
                index[4] = 1
                ws.cell(bagCount, 1).value = address
                ws.cell(bagCount, 2).value = privateKey.to_hex()
                bagCount = bagCount + 1
        elif address[len(address) - 5:len(address)] == "55555":
            print("privateKey:", privateKey)
            print("publicKey:", publicKey)
            print("address:", address)
            print("count:", count)
            if index[5] == 0:
                index[5] = 1
                ws.cell(bagCount, 1).value = address
                ws.cell(bagCount, 2).value = privateKey.to_hex()
                bagCount = bagCount + 1
        elif address[len(address) - 5:len(address)] == "66666":
            print("privateKey:", privateKey)
            print("publicKey:", publicKey)
            print("address:", address)
            print("count:", count)
            if index[6] == 0:
                index[6] = 1
                ws.cell(bagCount, 1).value = address
                ws.cell(bagCount, 2).value = privateKey.to_hex()
                bagCount = bagCount + 1
        elif address[len(address) - 5:len(address)] == "77777":
            print("privateKey:", privateKey)
            print("publicKey:", publicKey)
            print("address:", address)
            print("count:", count)
            if index[7] == 0:
                index[7] = 1
                ws.cell(bagCount, 1).value = address
                ws.cell(bagCount, 2).value = privateKey.to_hex()
                bagCount = bagCount + 1
        elif address[len(address) - 5:len(address)] == "88888":
            print("privateKey:", privateKey)
            print("publicKey:", publicKey)
            print("address:", address)
            print("count:", count)
            if index[8] == 0:
                index[8] = 1
                ws.cell(bagCount, 1).value = address
                ws.cell(bagCount, 2).value = privateKey.to_hex()
                bagCount = bagCount + 1
        elif address[len(address) - 5:len(address)] == "99999":
            print("privateKey:", privateKey)
            print("publicKey:", publicKey)
            print("address:", address)
            print("count:", count)
            if index[9] == 0:
                index[9] = 1
                ws.cell(bagCount, 1).value = address
                ws.cell(bagCount, 2).value = privateKey.to_hex()
                bagCount = bagCount + 1
        if index[1]==1 and index[2]==1 and index[3]==1 and index[4]==1 and index[5]==1 and index[6]==1 and index[7]==1 and index[8]==1 and index[9]==1:
            wb.save('wallet.xlsx')
            break
    return privateKey, publicKey, address


createNewETHWallet()